import openpyxl
from openpyxl.styles import Alignment, Border, Side
import json

# Ruta del archivo JSON
FILE_PATH = "participantes.json"
EXCEL_FILE_PATH = 'participantes.xlsx'

nombres_participantes = {}

class Participante:
    def __init__(self, id, riotId, discordName):
        self.id = int(id)
        self.riotId = riotId
        self.discordName = discordName
    def __str__(self) -> str:
        return f'{self.id} {self.riotId} {self.discordName}'
    def to_dict(self):
        return {
            "id": int(self.id),
            "riotId": self.riotId,
            "discordName": self.discordName
        }

def actualizar_excel():
    try:
        # Cargar el libro de trabajo
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
        ws = wb.active

        # Limpiar el contenido previo (opcional)
        ws.delete_rows(1, ws.max_row)

        # Agregar encabezados
        headers = ['ID', 'RiotID', 'Nombre Discord']
        ws.append(headers)

        # Agregar los datos de los participantes
        for participante in nombres_participantes.values():
            ws.append([str(participante.id), participante.riotId, participante.discordName])

        # Formatear la columna de ID como texto
        for cell in ws['A']:
            cell.number_format = '@'

        # Ajustar el ancho de las columnas A, B y C
        column_widths = {
            'A': 20,  # Ancho para la columna ID
            'B': 30,  # Ancho para la columna RiotID
            'C': 40   # Ancho para la columna Nombre Discord
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Definir el estilo de alineación y bordes
        alignment = Alignment(horizontal='center', vertical='center')
        border = Border(left=Side(border_style='thin'), 
                        right=Side(border_style='thin'), 
                        top=Side(border_style='thin'), 
                        bottom=Side(border_style='thin'))

        # Aplicar el estilo a todas las celdas
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = alignment
                cell.border = border

        # Guardar los cambios
        wb.save(EXCEL_FILE_PATH)
        print("Archivo Excel actualizado correctamente.")
    except Exception as e:
        print(f"Error al actualizar el archivo Excel: {e}")

def guardarIngreso(id, riotId, discordName):
    nombres_participantes[id] = Participante(id, riotId, discordName)
    guardar_participantes()
    actualizar_excel()
    print(nombres_participantes[id])
    
def verificarIngreso(id):
    return id in nombres_participantes

def guardar_participantes():
    with open(FILE_PATH, 'w') as file:
        json.dump({key: value.to_dict() for key, value in nombres_participantes.items()}, file)

# Función para cargar el diccionario desde un archivo JSON
def cargar_participantes():
    global nombres_participantes
    try:
        with open(FILE_PATH, 'r') as file:
            data = json.load(file)
            nombres_participantes = {int(key): Participante(id=int(key), riotId=value["riotId"], discordName=value["discordName"]) for key, value in data.items()}
    except FileNotFoundError:
        nombres_participantes = {}
  
def eliminarIngreso(id):
    if id in nombres_participantes:
        del nombres_participantes[id]
        guardar_participantes()  # Actualiza el JSON
        actualizar_excel()       # Actualiza el Excel
              
cargar_participantes()